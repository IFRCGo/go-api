from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from django.conf import settings
from openpyxl import load_workbook

XLSX_PATH = Path(settings.BASE_DIR) / "data" / "IFRC_Customs_Data.xlsx"
_cached_data = None

MASTER_SHEET = "DB_Master"


QA_REQUIRED_COLS = ["Section", "Question", "Country Response", "Notes/Comments"]


def _clean(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _find_header(ws, required_cols: List[str], max_rows: int = 50):
    req = set(required_cols)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        if not row:
            continue
        cols = {str(c).strip() for c in row if c is not None}
        if req.issubset(cols):
            return row_idx, row
    raise ValueError(f"Could not find header containing {required_cols} in sheet '{ws.title}'")


def _get_col_index(header_row) -> Dict[str, int]:
    return {str(name).strip(): idx for idx, name in enumerate(header_row) if name is not None}


def _ensure_country(countries: Dict, country: str):
    if country not in countries:
        countries[country] = {"country": country, "sections": {}}


def _add_item(countries: Dict, country: str, section: str, question: str, answer: str, notes: str = ""):
    _ensure_country(countries, country)
    if section not in countries[country]["sections"]:
        countries[country]["sections"][section] = []
    countries[country]["sections"][section].append({"question": question, "answer": answer, "notes": notes})


def _parse_db_master(ws, countries: Dict):
    # Find header row containing at least "Country name"
    header_row_idx, header = _find_header(ws, ["Country name"])
    col_index = _get_col_index(header)

    country_col = col_index.get("Country name")
    if country_col is None:
        raise ValueError("DB_Master is missing 'Country name' column")

    rows = ws.iter_rows(min_row=header_row_idx + 1, values_only=True)

    # Treat each non-empty cell as a Q&A item under a section
    default_section = "DB_Master"

    for r in rows:
        if not r:
            continue
        country = _clean(r[country_col] if country_col < len(r) else "")
        if not country:
            continue

        for col_name, idx in col_index.items():
            if col_name == "Country name":
                continue
            val = _clean(r[idx] if idx < len(r) else "")
            if val == "":
                continue
            _add_item(
                countries=countries,
                country=country,
                section=default_section,
                question=col_name,
                answer=val,
                notes="",
            )


def _parse_qa_sheet(ws, countries: Dict):
    header_row_idx, header = _find_header(ws, QA_REQUIRED_COLS)
    col_index = _get_col_index(header)

    i_section = col_index["Section"]
    i_question = col_index["Question"]
    i_answer = col_index["Country Response"]
    i_notes = col_index["Notes/Comments"]

    current_country: Optional[str] = None
    current_section: Optional[str] = None

    rows = ws.iter_rows(min_row=header_row_idx + 1, values_only=True)

    for r in rows:
        section = _clean(r[i_section] if i_section < len(r) else "")
        question = _clean(r[i_question] if i_question < len(r) else "")
        answer = _clean(r[i_answer] if i_answer < len(r) else "")
        notes = _clean(r[i_notes] if i_notes < len(r) else "")

        if section:
            current_section = section

        if not question:
            continue

        # Convention: a row whose question is "Country name:" sets current_country
        if question.lower().startswith("country name"):
            if answer:
                current_country = answer
                _ensure_country(countries, current_country)
            continue

        if not current_country:
            continue

        _add_item(
            countries=countries,
            country=current_country,
            section=current_section or ws.title,
            question=question,
            answer=answer,
            notes=notes,
        )


def load_customs_regulations() -> Dict:
    global _cached_data
    if _cached_data is not None:
        return _cached_data

    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Customs regulations XLSX not found at: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    countries: Dict[str, Dict] = {}

    # 1) Parse DB_Master if present
    if MASTER_SHEET in wb.sheetnames:
        _parse_db_master(wb[MASTER_SHEET], countries)

    # 2) Parse all other sheets with the consistent Q&A format
    for name in wb.sheetnames:
        if name == MASTER_SHEET:
            continue
        ws = wb[name]
        # Only parse if it looks like a Q&A sheet
        try:
            _find_header(ws, QA_REQUIRED_COLS, max_rows=30)
        except Exception:
            continue  # skip sheets that don't match the Q&A schema
        _parse_qa_sheet(ws, countries)

    result = {
        "metadata": {
            "source": "IFRC Customs & Import Regulations",
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "countries": [
            {
                "country": c["country"],
                "sections": [{"section": section, "items": items} for section, items in c["sections"].items()],
            }
            for c in countries.values()
        ],
    }

    _cached_data = result
    return result
