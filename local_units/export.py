from pathlib import Path
from typing import Any, Callable

from django.conf import settings
from django.http import HttpResponse
from openpyxl import load_workbook

from local_units.bulk_upload import BASE_HEADER_MAP, HEALTH_HEADER_MAP, START_ROW
from local_units.models import HealthData, LocalUnit

LOCAL_UNIT_EXPORT_HANDLERS: dict[str, Callable[[LocalUnit], Any]] = {
    "date_of_data": lambda u: u.date_of_data.strftime("%Y-%m-%d") if u.date_of_data else "",
    "visibility": lambda u: u.get_visibility_display() if u.visibility else "",
    "level": lambda u: u.level.name if u.level else "",
    "latitude": lambda u: u.location.y if u.location else "",
    "longitude": lambda u: u.location.x if u.location else "",
}

HEALTH_EXPORT_HANDLERS: dict[str, Callable[[HealthData], Any]] = {
    "affiliation": lambda h: h.affiliation.name if h.affiliation else "",
    "functionality": lambda h: h.functionality.name if h.functionality else "",
    "health_facility_type": lambda h: h.health_facility_type.name if h.health_facility_type else "",
    "hospital_type": lambda h: h.hospital_type.name if h.hospital_type else "",
    "primary_health_care_center": lambda h: h.primary_health_care_center.name if h.primary_health_care_center else "",
    "general_medical_services": lambda h: ", ".join(m.name for m in h.general_medical_services.all()),
    "specialized_medical_beyond_primary_level": lambda h: ", ".join(
        m.name for m in h.specialized_medical_beyond_primary_level.all()
    ),
    "blood_services": lambda h: ", ".join(m.name for m in h.blood_services.all()),
    "professional_training_facilities": lambda h: ", ".join(m.name for m in h.professional_training_facilities.all()),
    "other_profiles": lambda h: ", ".join(p.position for p in h.other_profiles.all()),
    "other_medical_heal": lambda h: "Yes" if h.other_medical_heal else "No",
    "is_in_patient_capacity": lambda h: "Yes" if h.is_in_patient_capacity else "No",
    "is_teaching_hospital": lambda h: "Yes" if h.is_teaching_hospital else "No",
    "is_isolation_rooms_wards": lambda h: "Yes" if h.is_isolation_rooms_wards else "No",
    "is_warehousing": lambda h: "Yes" if h.is_warehousing else "No",
    "is_cold_chain": lambda h: "Yes" if h.is_cold_chain else "No",
}


def _resolve_value(unit: LocalUnit, field: str, is_health: bool):
    if field in LOCAL_UNIT_EXPORT_HANDLERS:
        return LOCAL_UNIT_EXPORT_HANDLERS[field](unit)

    # First check if there's a specific handler for health fields and fall back to direct attribute access
    if is_health and unit.health and field in HEALTH_EXPORT_HANDLERS:
        return HEALTH_EXPORT_HANDLERS[field](unit.health)

    if is_health and unit.health and hasattr(unit.health, field):
        return getattr(unit.health, field) or ""

    return getattr(unit, field, "") or ""


def export_local_units_to_excel(
    queryset,
    *,
    is_health: bool = False,
    file_name: str | None = None,
):
    """
    Exports local units to an Excel file and returns it as an HTTP response.
    """

    header_map = HEALTH_HEADER_MAP if is_health else BASE_HEADER_MAP

    if is_health:
        template = "Health-Care-Bulk-Import-Template-Local-Units.xlsm"
    else:
        template = "Administrative-Bulk-Import-Template-Local-Units.xlsx"

    template_path = Path(settings.BASE_DIR) / "go-static/files/local_units" / template

    if not template_path.exists():
        raise FileNotFoundError(f"Excel template not found: {template_path}")

    wb = load_workbook(template_path, read_only=False)
    ws = wb.active

    if ws is None:
        raise ValueError("Worksheet could not be loaded from template.")

    headers = [cell.value for cell in ws[2] if cell.value]
    start_row = START_ROW

    for row_idx, unit in enumerate(queryset, start=start_row):
        for col_idx, header in enumerate(headers, start=1):
            field = header_map.get(str(header).strip())
            if not field:
                continue
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=_resolve_value(unit, field, is_health),
            )

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{file_name or "local_units_export.xlsx"}"'

    wb.save(response)
    return response
