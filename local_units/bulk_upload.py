import io
import logging
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, Generic, Literal, Optional, Type, TypeVar

import openpyxl
from django.core.files.base import ContentFile
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from rest_framework.exceptions import ErrorDetail
from rest_framework.serializers import Serializer

from local_units.models import HealthData, LocalUnit, LocalUnitBulkUpload, LocalUnitType
from local_units.utils import get_model_field_names
from main.managers import BulkCreateManager

logger = logging.getLogger(__name__)

ContextType = TypeVar("ContextType")


class BulkUploadError(Exception):
    """Custom exception for bulk upload errors."""

    pass


class ErrorWriter:
    ERROR_ROW_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    def __init__(self, fieldnames: list[str], header_map: dict[str, str]):
        """Initialize workbook and header row."""
        self._header_map = header_map or {}
        self._reverse_header_map = {v: k for k, v in self._header_map.items()}
        # Convert model field names to xlsx template headers
        display_header = [self._reverse_header_map.get(field, field) for field in fieldnames]

        self._fieldnames = ["Upload Status"] + display_header
        self._workbook = Workbook()
        self._ws = self._workbook.active

        self._ws.append(self._fieldnames)
        self._existing_error_columns = set()
        self._rows: list[dict[str, str]] = []
        self._has_errors = False

    def _format_errors(self, errors: dict) -> dict[str, list[str]]:
        """Recursively flatten DRF errors."""
        formatted = {}
        for key, value in errors.items():
            if isinstance(value, dict):
                formatted.update(self._format_errors(value))
            elif isinstance(value, list):
                header = self._reverse_header_map.get(key, key)
                formatted[header] = [self._clean_message(v) for v in value]
            else:
                header = self._reverse_header_map.get(key, key)
                formatted[header] = [self._clean_message(value)]
        return formatted

    def _clean_message(self, msg: any) -> str:
        if isinstance(msg, ErrorDetail):
            return str(msg)
        return str(msg)

    def _add_error_columns(self, fields: list[str]):
        """Ensure field has a matching _error column."""
        for field in fields:
            col_name = f"{field}_error"
            if col_name in self._existing_error_columns:
                continue
            self._existing_error_columns.add(col_name)

            if field in self._fieldnames:
                idx = self._fieldnames.index(field) + 1
                self._fieldnames.insert(idx, col_name)
                self._ws.insert_cols(idx + 1)
            else:
                self._fieldnames.append(col_name)
        for i, col_name in enumerate(self._fieldnames, start=1):
            self._ws.cell(row=1, column=i, value=col_name)

    def write(
        self,
        row: dict[str, any],
        status: Literal[LocalUnitBulkUpload.Status.SUCCESS, LocalUnitBulkUpload.Status.FAILED],
        error_detail: dict | None = None,
    ):
        row_display = {self._reverse_header_map.get(k, k): v for k, v in row.items()}
        row_out = {col: row_display.get(col, "") for col in self._fieldnames}
        row_out["Upload Status"] = status.name

        if status == LocalUnitBulkUpload.Status.FAILED and error_detail:
            formatted = self._format_errors(error_detail)
            self._add_error_columns(list(formatted.keys()))
            for field, msgs in formatted.items():
                row_out[f"{field}_error"] = "; ".join(msgs)
            self._has_errors = True

        self._ws.append([row_out.get(col, "") for col in self._fieldnames])

        if status == LocalUnitBulkUpload.Status.FAILED:
            for cell in self._ws[self._ws.max_row]:
                cell.fill = self.ERROR_ROW_FILL

    def to_content_file(self) -> ContentFile:
        """Export workbook as Content File for Django."""
        buffer = io.BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        return ContentFile(buffer.getvalue())


class BaseBulkUpload(Generic[ContextType]):
    serializer_class: Type[Serializer]
    HEADER_MAP: dict[str, str]

    def __init__(self, bulk_upload: LocalUnitBulkUpload):
        if self.serializer_class is None:
            raise ValueError("serializer_class must be defined in the subclass.")

        self.bulk_upload = bulk_upload
        self.success_count: int = 0
        self.failed_count: int = 0
        self.bulk_manager: BulkCreateManager = BulkCreateManager(chunk_size=500)
        self.error_writer: Optional[ErrorWriter] = None

    def get_context(self) -> ContextType:
        raise NotImplementedError("Subclasses must implement get_context method.")

    def delete_existing_local_unit(self):
        """Delete existing local units based on the context."""
        pass

    def _validate_type(self, fieldnames) -> None:
        pass

    def is_excel_data_empty(self, sheet, data_start_row=4):
        """Check if file is empty or not"""
        for row in sheet.iter_rows(values_only=True, min_row=data_start_row):
            if any(cell is not None for cell in row):
                return False
        return True

    def process_row(self, data: Dict[str, Any]) -> bool:
        serializer = self.serializer_class(data=data)
        if serializer.is_valid():
            self.bulk_manager.add(LocalUnit(**serializer.validated_data))
            return True
        self.error_detail = serializer.errors
        return False

    def run(self) -> None:
        header_row_index = 2
        data_row_index = header_row_index + 2

        try:
            with self.bulk_upload.file.open("rb") as f:
                workbook = openpyxl.load_workbook(f, data_only=True, read_only=True)
                sheet = workbook.active

                headers = next(sheet.iter_rows(values_only=True, min_row=header_row_index, max_row=header_row_index))
                raw_fieldnames = [str(h).strip() for h in headers if h and str(h).strip()]
                header_map = self.HEADER_MAP or {}
                mapped_fieldnames = [header_map.get(h, h) for h in raw_fieldnames]
                fieldnames = mapped_fieldnames

                if self.is_excel_data_empty(sheet, data_start_row=data_row_index):
                    raise BulkUploadError("The uploaded file is empty. Please provide at least one data row.")

                self._validate_type(fieldnames)
                self.error_writer = ErrorWriter(fieldnames=raw_fieldnames, header_map=header_map)
                context = self.get_context().__dict__

                with transaction.atomic():
                    self.delete_existing_local_unit()

                    for row_index, row_values in enumerate(
                        sheet.iter_rows(values_only=True, min_row=data_row_index),
                        start=data_row_index,
                    ):
                        if not any(cell is not None for cell in row_values):
                            continue  # skip empty rows

                        row_dict = dict(zip(fieldnames, row_values))
                        row_dict = {**row_dict, **context}

                        # Convert date/datetime to str
                        for key, value in row_dict.items():
                            if isinstance(value, (datetime, date)):
                                row_dict[key] = value.strftime("%Y-%m-%d")

                        if self.process_row(row_dict):
                            self.success_count += 1
                            self.error_writer.write(row_dict, status=LocalUnitBulkUpload.Status.SUCCESS)
                        else:
                            self.failed_count += 1
                            self.error_writer.write(
                                row_dict,
                                status=LocalUnitBulkUpload.Status.FAILED,
                                error_detail=self.error_detail,
                            )
                            logger.warning(f"[BulkUpload:{self.bulk_upload.pk}] Row {row_index} failed")

                    if self.failed_count > 0:
                        raise BulkUploadError()

                    self.bulk_manager.done()
                    self._finalize_success()

                workbook.close()

        except Exception as e:
            self.bulk_upload.status = LocalUnitBulkUpload.Status.FAILED
            self.bulk_upload.error_message = str(e)
            self.bulk_upload.save(update_fields=["status", "error_message"])
            if isinstance(e, BulkUploadError):
                logger.warning(f"[BulkUpload:{self.bulk_upload.pk}]  error: {e}")
            self._finalize_failure()

    def _finalize_success(self) -> None:
        self.bulk_upload.success_count = self.success_count
        self.bulk_upload.failed_count = self.failed_count
        self.bulk_upload.status = LocalUnitBulkUpload.Status.SUCCESS
        self.bulk_upload.save(update_fields=["success_count", "failed_count", "status"])

        logger.info(f"[BulkUpload:{self.bulk_upload.pk}] SUCCESS: {self.success_count} rows.")

    def _finalize_failure(self) -> None:
        if self.error_writer:
            error_file = self.error_writer.to_content_file()
            self.bulk_upload.error_file.save("error_file.xlsx", error_file, save=True)

        self.bulk_upload.success_count = self.success_count
        self.bulk_upload.failed_count = self.failed_count
        self.bulk_upload.status = LocalUnitBulkUpload.Status.FAILED
        self.bulk_upload.save(update_fields=["success_count", "failed_count", "status", "error_file"])

        logger.info(f"[BulkUpload:{self.bulk_upload.pk}] FAILED: " f"{self.success_count} succeeded, {self.failed_count} failed.")


@dataclass(frozen=True)
class LocalUnitUploadContext:
    country: int
    type: int
    created_by: int
    bulk_upload: int


class BaseBulkUploadLocalUnit(BaseBulkUpload[LocalUnitUploadContext]):
    HEADER_MAP = {
        "Date of Update": "date_of_data",
        "Local Unit Name (En)": "english_branch_name",
        "Local Unit Name (Local)": "local_branch_name",
        "Visibility": "visibility",
        "Coverage": "level",
        "Sub-type": "subtype",
        "Focal Person (En)": "focal_person_en",
        "Source (En)": "source_en",
        "Source (Local)": "source_loc",
        "Focal Person (Local)": "focal_person_loc",
        "Address (Local)": "address_loc",
        "Address (En)": "address_en",
        "Locality (Local)": "city_loc",
        "Locality (En)": "city_en",
        "Local Unit Post Code": "postcode",
        "Local Unit Email": "email",
        "Local Unit Phone Number": "phone",
        "Local Unit Website": "link",
        "Latitude": "latitude",
        "Longitude": "longitude",
    }

    def __init__(self, bulk_upload: LocalUnitBulkUpload):
        from local_units.serializers import LocalUnitBulkUploadDetailSerializer

        self.serializer_class = LocalUnitBulkUploadDetailSerializer
        super().__init__(bulk_upload)

    def get_context(self) -> LocalUnitUploadContext:
        return LocalUnitUploadContext(
            country=self.bulk_upload.country_id,
            type=self.bulk_upload.local_unit_type_id,
            created_by=self.bulk_upload.triggered_by_id,
            bulk_upload=self.bulk_upload.pk,
        )

    def delete_existing_local_unit(self):
        queryset = LocalUnit.objects.filter(
            country_id=self.bulk_upload.country_id,
            type_id=self.bulk_upload.local_unit_type_id,
        )
        if queryset.exists():
            count, _ = queryset.delete()
            logger.info(f"Deleted {count} existing local units before upload.")
        else:
            logger.info("No existing local units found for deletion.")

    def _validate_type(self, fieldnames: list[str]) -> None:

        health_field_names = set(get_model_field_names(HealthData))
        present_health_fields = {h for h in health_field_names if h.lower() in [f.lower() for f in fieldnames]}

        local_unit_type = LocalUnitType.objects.filter(id=self.bulk_upload.local_unit_type_id).first()
        if not local_unit_type:
            raise ValueError("Invalid local unit type")

        if present_health_fields and local_unit_type.name.strip().lower() != "health care":
            raise BulkUploadError(
                f"You cannot upload Healthcare data when the Local Unit type is set to '{local_unit_type.name}'."
            )


class BulkUploadHealthData(BaseBulkUpload[LocalUnitUploadContext]):
    # Local Unit headers + Health Data headers
    HEADER_MAP = {
        **BaseBulkUploadLocalUnit.HEADER_MAP,
        **{
            "Focal Person Email": "focal_point_email",
            "Focal Person Phone Number": "focal_point_phone_number",
            "Focal Person Position": "focal_point_position",
            "Health Facility Type": "health_facility_type",
            "Other Facility Type": "other_facility_type",
            "Affiliation": "affiliation",
            "Other Affiliation": "other_affiliation",
            "Functionality": "functionality",
            "Primary Health Care Center": "primary_health_care_center",
            "Specialities": "speciality",
            "Hospital Type": "hospital_type",
            "Teaching Hospital": "is_teaching_hospital",
            "In-patient Capacity": "is_in_patient_capacity",
            "Isolation Rooms": "is_isolation_rooms_wards",
            "Number of Isolation Beds": "number_of_isolation_rooms",
            "Warehousing": "is_warehousing",
            "Cold Chain": "is_cold_chain",
            "Maximum Bed Capacity": "maximum_capacity",
            "General Medical Services": "general_medical_services",
            "Specialized Medical Services (beyond primary level)": "specialized_medical_beyond_primary_level",
            "Blood Services": "blood_services",
            "Other Services": "other_services",
            "Total Number of Human Resources": "total_number_of_human_resource",
            "General Practitioners": "general_practitioner",
            "Resident Doctors": "residents_doctor",
            "Specialists": "specialist",
            "Nurses": "nurse",
            "Nursing Aids": "nursing_aid",
            "Dentists": "dentist",
            "Midwife": "midwife",
            "Pharmacists": "pharmacists",
            "Other Profiles": "other_profiles",
            "Other Training Facility": "other_training_facilities",
            "Professional Training Facilities": "professional_training_facilities",
            "Ambulance Type A": "ambulance_type_a",
            "Ambulance Type B": "ambulance_type_b",
            "Ambulance Type C": "ambulance_type_c",
        },
    }

    def __init__(self, bulk_upload: LocalUnitBulkUpload):
        from local_units.serializers import LocalUnitBulkUploadDetailSerializer

        self.serializer_class = LocalUnitBulkUploadDetailSerializer
        super().__init__(bulk_upload)

        self.health_field_names = get_model_field_names(
            HealthData,
        )

    def get_context(self) -> LocalUnitUploadContext:
        return LocalUnitUploadContext(
            country=self.bulk_upload.country_id,
            type=self.bulk_upload.local_unit_type_id,
            created_by=self.bulk_upload.triggered_by_id,
            bulk_upload=self.bulk_upload.pk,
        )

    def delete_existing_local_unit(self):
        local_unit_queryset = LocalUnit.objects.filter(
            country_id=self.bulk_upload.country_id,
            type_id=self.bulk_upload.local_unit_type_id,
        )
        if local_unit_queryset.exists():
            health_data_ids = local_unit_queryset.filter(health__isnull=False).values_list("health__id", flat=True)
            health_data_count, _ = HealthData.objects.filter(id__in=health_data_ids).delete()
            count, _ = local_unit_queryset.delete()
            logger.info(f"Deleted {count} existing local units and {health_data_count} health data before upload.")
        else:
            logger.info("No existing local units found for deletion.")

    def process_row(self, data: dict[str, any]) -> bool:
        from local_units.serializers import HealthDataBulkUploadSerializer

        health_data = {k: data.get(k) for k in data.keys() if k in self.health_field_names}

        if health_data:
            health_serializer = HealthDataBulkUploadSerializer(data=health_data)
            if not health_serializer.is_valid():
                self.error_detail = health_serializer.errors
                return False
            health_instance = health_serializer.save()
            for k in health_data.keys():
                data.pop(k, None)
            data["health"] = health_instance.pk
            return super().process_row(data)
