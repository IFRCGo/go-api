from datetime import datetime

import pytz
from django.conf import settings
from django.db import transaction
from django.db.models import Count, F, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.translation import get_language as django_get_language
from django_filters import rest_framework as filters
from django_filters.widgets import CSVWidget
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from rest_framework import mixins, permissions, response
from rest_framework import status as drf_status
from rest_framework import views, viewsets
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.settings import api_settings

from api.models import Country, Region
from deployments.models import SectorTag
from main.permissions import DenyGuestUserMutationPermission, DenyGuestUserPermission
from main.utils import SpreadSheetContentNegotiation
from per.cache import OpslearningSummaryCacheHelper
from per.filter_set import (
    PerDocumentFilter,
    PerOverviewFilter,
    PerPrioritizationFilter,
    PerWorkPlanFilter,
)
from per.permissions import (
    OpsLearningPermission,
    PerDocumentUploadPermission,
    PerGeneralPermission,
    PerPermission,
)
from per.task import generate_summary
from per.utils import filter_per_queryset_by_user_access

from .admin_classes import RegionRestrictedAdmin
from .custom_renderers import NarrowCSVRenderer
from .models import (
    AreaResponse,
    AssessmentType,
    FormAnswer,
    FormArea,
    FormComponent,
    FormComponentQuestionAndAnswer,
    FormComponentResponse,
    FormData,
    FormPrioritization,
    FormPrioritizationComponent,
    FormQuestion,
    FormQuestionGroup,
    NiceDocument,
    OpsLearning,
    OpsLearningCacheResponse,
    OpsLearningComponentCacheResponse,
    OpsLearningSectorCacheResponse,
    OrganizationTypes,
    Overview,
    PerAssessment,
    PerComponentRating,
    PerDocumentUpload,
    PerFile,
    PerWorkPlan,
)
from .serializers import (
    FormAnswerSerializer,
    FormAreaSerializer,
    FormComponentSerializer,
    FormPrioritizationSerializer,
    FormQuestionGroupSerializer,
    FormQuestionSerializer,
    LatestCountryOverviewSerializer,
    ListNiceDocSerializer,
    NiceDocumentSerializer,
    OpsLearningCSVSerializer,
    OpsLearningInSerializer,
    OpsLearningOrganizationTypeSerializer,
    OpsLearningSerializer,
    OpsLearningStatSerializer,
    OpsLearningSummarySerializer,
    PerAssessmentSerializer,
    PerDocumentUploadSerializer,
    PerFileInputSerializer,
    PerFileSerializer,
    PerFormDataSerializer,
    PerOptionsSerializer,
    PerOverviewSerializer,
    PerProcessSerializer,
    PerWorkPlanSerializer,
    PublicOpsLearningSerializer,
    PublicPerAssessmentSerializer,
    PublicPerCountrySerializer,
    PublicPerProcessSerializer,
    UserPerCountrySerializer,
)

# Helpers for transformed "-2" endpoints
AREA_NAMES = {
    1: "Policy Strategy and Standards",
    2: "Analysis and planning",
    3: "Operational capacity",
    4: "Coordination",
    5: "Operations support",
}

AFFIRMATIVE_WORDS = {"yes", "si", "sí", "oui", "da", "ja", "sim", "aye", "yep", "igen", "hai", "evet", "是", "はい", "예", "نعم"}


def _contains_affirmative(text: str) -> bool:
    if not text or not isinstance(text, str):
        return False
    try:
        import unicodedata

        normalized = unicodedata.normalize("NFD", text.lower())
        normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    except Exception:
        normalized = text.lower()
    return any(word in normalized for word in AFFIRMATIVE_WORDS)


class PERDocsFilter(filters.FilterSet):
    id = filters.NumberFilter(field_name="id", lookup_expr="exact")

    class Meta:
        model = NiceDocument
        fields = {
            "id": ("exact",),
        }


class PERDocsViewset(viewsets.ReadOnlyModelViewSet):
    """To collect PER Documents"""

    # Duplicate of FormDataViewset
    queryset = NiceDocument.objects.all()
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    get_request_user_regions = RegionRestrictedAdmin.get_request_user_regions
    get_filtered_queryset = RegionRestrictedAdmin.get_filtered_queryset
    filterset_class = PERDocsFilter

    def get_queryset(self):
        queryset = NiceDocument.objects.all()
        cond1 = Q()
        cond2 = Q()

        cond3 = Q()
        if "new" in self.request.query_params.keys():
            last_duedate = settings.PER_LAST_DUEDATE
            tmz = pytz.timezone("Europe/Zurich")
            if not last_duedate:
                last_duedate = tmz.localize(datetime(2000, 11, 15, 9, 59, 25, 0))
            cond1 = Q(created_at__gt=last_duedate)
        if "country" in self.request.query_params.keys():
            cid = self.request.query_params.get("country", None) or 0
            country = Country.objects.filter(pk=cid)
            if country:
                cond2 = Q(country_id=country[0].id)
        if "visible" in self.request.query_params.keys():
            cond3 = Q(visibility=1)
        queryset = NiceDocument.objects.filter(cond1 & cond2 & cond3)
        if queryset.exists():
            queryset = self.get_filtered_queryset(self.request, queryset, 4)
        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return ListNiceDocSerializer
        return NiceDocumentSerializer


class FormAreaFilter(filters.FilterSet):
    id = filters.NumberFilter(field_name="id", lookup_expr="exact")
    area_num = filters.NumberFilter(field_name="area_num", lookup_expr="exact")

    class Meta:
        model = FormArea
        fields = {"id": ("exact",), "area_num": ("exact",)}


class FormAreaViewset(viewsets.ReadOnlyModelViewSet):
    """PER Form Areas Viewset"""

    serializer_class = FormAreaSerializer
    queryset = FormArea.objects.all().order_by("area_num")
    filterset_class = FormAreaFilter


class FormComponentFilter(filters.FilterSet):
    area_id = filters.NumberFilter(field_name="area__id", lookup_expr="exact")
    exclude_subcomponents = filters.BooleanFilter(
        method="get_exclude_subcomponents",
    )

    class Meta:
        model = FormComponent
        fields = {"area": ("exact",)}

    def get_exclude_subcomponents(self, queryset, name, value):
        if value:
            return queryset.exclude(component_num=14, is_parent__isnull=True)
        return queryset


class FormComponentViewset(viewsets.ReadOnlyModelViewSet):
    """PER Form Components Viewset"""

    serializer_class = FormComponentSerializer
    filterset_class = FormComponentFilter

    def get_queryset(self):
        return FormComponent.objects.all().order_by("area__area_num", "component_num", "component_letter").select_related("area")


class FormQuestionFilter(filters.FilterSet):
    area_id = filters.NumberFilter(field_name="component__area__id", lookup_expr="exact")

    class Meta:
        model = FormQuestion
        fields = {"component": ("exact",)}


class FormQuestionViewset(viewsets.ReadOnlyModelViewSet):
    """PER Form Questions Viewset"""

    serializer_class = FormQuestionSerializer
    filterset_class = FormQuestionFilter
    ordering_fields = "__all__"

    def get_queryset(self):
        return (
            FormQuestion.objects.all()
            .order_by("component__component_num", "question_num", "question")
            .select_related("component", "component__area")
            .prefetch_related("answers")
        )


class FormQuestionGroupViewset(viewsets.ReadOnlyModelViewSet):
    """PER From Question Group ViewSet"""

    serializer_class = FormQuestionGroupSerializer

    def get_queryset(self):
        return FormQuestionGroup.objects.select_related("component")


class FormAnswerViewset(viewsets.ReadOnlyModelViewSet):
    """PER Form Answers Viewset"""

    serializer_class = FormAnswerSerializer
    queryset = FormAnswer.objects.all()
    ordering_fields = "__all__"


class CountryPublicPerStatsViewset(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = LatestCountryOverviewSerializer
    filterset_class = PerOverviewFilter

    def get_queryset(self):
        return Overview.objects.select_related("country", "type_of_assessment").order_by("-created_at")


class CountryPerStatsViewset(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = LatestCountryOverviewSerializer
    filterset_class = PerOverviewFilter
    permission_classes = [IsAuthenticated, DenyGuestUserPermission]

    def get_queryset(self):
        return Overview.objects.select_related("country", "type_of_assessment").order_by("-created_at")


class PerOverviewViewSet(viewsets.ModelViewSet):
    serializer_class = PerOverviewSerializer
    permission_classes = [IsAuthenticated, PerPermission, DenyGuestUserPermission]
    filterset_class = PerOverviewFilter
    ordering_fields = "__all__"
    get_request_user_regions = RegionRestrictedAdmin.get_request_user_regions
    get_filtered_queryset = RegionRestrictedAdmin.get_filtered_queryset

    def get_queryset(self):
        queryset = Overview.objects.select_related("country", "user")
        return self.get_filtered_queryset(self.request, queryset, dispatch=0)


class ExportPerView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, DenyGuestUserPermission]

    content_negotiation_class = SpreadSheetContentNegotiation

    def get(self, request, pk, format=None):
        per = get_object_or_404(Overview, pk=pk)
        per_queryset = Overview.objects.filter(id=per.id)
        # if per.extracted_at is None or per.updated_at > per.extracted_at:
        wb = Workbook()
        ws = wb.active
        ws.title = "Overview"
        # Overview Columns
        ws.row_dimensions[1].height = 70
        overview_columns = [
            "National Society",
            "Date of creation (register)",
            "Date of last update (of the whole process)",
            "Date of Orientation",
            "Orientation document uploaded? (Yes/No)",
            "Date of Current PER Assessment",
            "Type of Assessment",
            "Branches involved",
            "Method",
            "Epidemic Considerations",
            "Urban Considerations",
            "Climate and env considerations",
            "Migration Considerations",
            "PER process cycle",
            "Work-plan development date planned",
            "Work-plan revision date planned",
            "NS FP name",
            "NS FP email",
            "NS FP phone number",
            "NS Second FP name",
            "NS Second FP email",
            "NS Second FP phone number",
            "Partner FP name",
            "Partner FP email",
            "Partner FP phone number",
            "Partner FP organization",
            "PER facilitator name",
            "PER facilitator email",
            "PER facilitator phone number",
            "PER facilitator other contact",
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(overview_columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
        for per in per_queryset:
            row_num += 1

            overview_rows = [
                per.country.name,
                per.created_at.date(),
                per.updated_at.date(),
                per.date_of_orientation,
                "Yes" if len(per.orientation_documents.all()) else "No",
                per.date_of_assessment,
                per.type_of_assessment.name,
                per.branches_involved,
                per.get_assessment_method_display(),
                per.assess_preparedness_of_country,
                per.assess_urban_aspect_of_country,
                per.assess_climate_environment_of_country,
                per.assess_migration_aspect_of_country,
                per.assessment_number,
                per.workplan_development_date,
                per.workplan_revision_date,
                per.ns_focal_point_name,
                per.ns_focal_point_email,
                per.ns_focal_point_phone,
                per.ns_second_focal_point_name,
                per.ns_second_focal_point_email,
                per.ns_second_focal_point_phone,
                per.partner_focal_point_name,
                per.partner_focal_point_email,
                per.partner_focal_point_phone,
                per.partner_focal_point_organization,
                per.facilitator_name,
                per.facilitator_email,
                per.facilitator_phone,
                per.facilitator_contact,
            ]

            for col_num, cell_value in enumerate(overview_rows, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Assessment
        ws_assessment = wb.create_sheet("Assessment")
        ws_assessment.row_dimensions[1].height = 70
        assessment_columns = [
            "Component number",
            "Component letter",
            "Component description",
            "Benchmark number",
            "Benchmark descprition",
            "Benchmark answer (Yes/No/Partially)",
            "Benchmark notes",
            "Consideration notes epidemic",
            "Consideration notes urban",
            "Consideration notes climate",
            "Consideration notes migration",
            "Component rating",
            "Component notes",
        ]
        assessment_num = 1
        for col_num, column_title in enumerate(assessment_columns, 1):
            cell = ws_assessment.cell(row=assessment_num, column=col_num)
            cell.value = column_title

        assessment_rows = []
        assessment_queryset = (
            PerAssessment.objects.filter(overview=per.id)
            .order_by("area_responses__component_response__component__component_num")
            .prefetch_related(
                Prefetch(
                    "area_responses",
                    queryset=AreaResponse.objects.filter(perassessment__overview=per.id).prefetch_related(
                        Prefetch(
                            "component_response",
                            queryset=FormComponentResponse.objects.filter(arearesponse__perassessment__overview=per.id)
                            .exclude(component_id=14)
                            .prefetch_related(
                                Prefetch(
                                    "question_responses",
                                    queryset=FormComponentQuestionAndAnswer.objects.filter(
                                        formcomponentresponse__arearesponse__perassessment__overview=per.id
                                    ),
                                )
                            ),
                        )
                    ),
                )
            )
        )
        if assessent := assessment_queryset.first():
            for area_response in assessent.area_responses.all():
                for co in area_response.component_response.all():
                    question_answer = co.question_responses.all()
                    for question in question_answer:
                        assessment_inner = [
                            co.component.component_num,
                            co.component.component_letter,
                            co.component.description_en,
                            (
                                str(question.question.component.component_num) + "." + str(question.question.question_num)
                                if question.question
                                else None
                            ),
                            question.question.question if question.question else None,
                            question.answer.text if question.answer else None,
                            question.notes,
                            co.epi_considerations,
                            co.urban_considerations,
                            co.climate_environmental_considerations,
                            co.migration_considerations,
                            co.rating.title if co.rating else None,
                            co.notes,
                        ]
                        assessment_rows.append(assessment_inner)

        for row_num, row_data in enumerate(assessment_rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws_assessment.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Prioritization
        ws_prioritization = wb.create_sheet("Prioritization")
        ws_prioritization.row_dimensions[1].height = 70
        prioritization_columns = [
            "Prioritized component number",
            "Prioritized component letter",
            "Prioritized component description",
            "Justification",
        ]
        prioritization_rows = []
        prioritization_num = 1
        for col_num, column_title in enumerate(prioritization_columns, 1):
            cell = ws_prioritization.cell(row=prioritization_num, column=col_num)
            cell.value = column_title

        prioritization_queryset = (
            FormPrioritizationComponent.objects.filter(
                formprioritization__overview=per.id,
            )
            .order_by("component__component_num")
            .exclude(component_id=14)
        )
        for prioritization in prioritization_queryset:
            prioritization_inner = [
                prioritization.component.component_num,
                prioritization.component.component_letter,
                prioritization.component.description,
                prioritization.justification_text,
            ]
            prioritization_rows.append(prioritization_inner)
        for row_num, row_data in enumerate(prioritization_rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws_prioritization.cell(row=row_num, column=col_num)
                cell.value = cell_value
        # Workplan
        ws_workplan = wb.create_sheet("Workplan")
        ws_workplan.row_dimensions[1].height = 70
        workplan_columns = [
            "Actions",
            "Number of component related",
            "Letter of component related",
            "Description of component related",
            "Due date",
            "Supported by",
            "Supporting National Society",
            "Status",
        ]
        workplan_rows = []
        workplan_num = 1
        for col_num, column_title in enumerate(workplan_columns, 1):
            cell = ws_workplan.cell(row=workplan_num, column=col_num)
            cell.value = column_title

        workplan_queryset = PerWorkPlan.objects.filter(overview=per.id)
        if workplan_queryset.exists():
            for workplan in workplan_queryset.first().prioritized_action_responses.all():
                workplan_inner = [
                    workplan.actions,
                    workplan.component.component_num,
                    workplan.component.component_letter,
                    workplan.component.description_en,
                    workplan.due_date,
                    workplan.get_supported_by_organization_type_display(),
                    workplan.supported_by.name if workplan.supported_by else None,
                    workplan.get_status_display(),
                ]
                workplan_rows.append(workplan_inner)
        if workplan_queryset.exists():
            for workplan in workplan_queryset.first().additional_action_responses.all():
                workplan_inner = [
                    workplan.actions,
                    None,
                    None,
                    None,
                    workplan.due_date,
                    workplan.get_supported_by_organization_type_display(),
                    workplan.supported_by.name if workplan.supported_by else None,
                    workplan.get_status_display(),
                ]
                workplan_rows.append(workplan_inner)
        for row_num, row_data in enumerate(workplan_rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws_workplan.cell(row=row_num, column=col_num)
                cell.value = cell_value

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=export.xlsx"
        wb.save(response)
        return response


class NewPerWorkPlanViewSet(viewsets.ModelViewSet):
    permission_classes = (IsAuthenticated, PerGeneralPermission, DenyGuestUserPermission)
    queryset = PerWorkPlan.objects.all()
    serializer_class = PerWorkPlanSerializer
    filterset_class = PerWorkPlanFilter
    ordering_fields = "__all__"


class PerFormDataViewSet(viewsets.ModelViewSet):
    serializer_class = PerFormDataSerializer
    queryset = FormData.objects.all()
    ordering_fields = "__all__"


class FormPrioritizationViewSet(viewsets.ModelViewSet):
    serializer_class = FormPrioritizationSerializer
    queryset = FormPrioritization.objects.all()
    filterset_class = PerPrioritizationFilter
    permission_classes = (IsAuthenticated, PerGeneralPermission, DenyGuestUserPermission)
    ordering_fields = "__all__"


class PublicFormPrioritizationViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = FormPrioritizationSerializer
    queryset = FormPrioritization.objects.all()
    ordering_fields = "__all__"


class PublicFormPrioritizationViewSet2(PublicFormPrioritizationViewSet):
    """Adds simplified prioritized components array (mirrors JS)."""

    def list(self, request, *args, **kwargs):
        resp = super().list(request, *args, **kwargs)
        data = resp.data
        results = data.get("results") if isinstance(data, dict) else data
        if results:
            for item in results:
                pcs = []
                for resp_item in item.get("prioritized_action_responses", []) or []:
                    cd = resp_item.get("component_details") or {}
                    pcs.append(
                        {
                            "componentId": resp_item.get("component"),
                            "componentTitle": cd.get("title"),
                            "areaTitle": (
                                cd.get("area_title") or AREA_NAMES.get(cd.get("area"))
                                if isinstance(cd.get("area"), int)
                                else cd.get("area_title")
                            ),
                            "description": cd.get("description"),
                        }
                    )
                item["components"] = pcs
        return Response(data)


class PerOptionsView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, DenyGuestUserPermission]
    ordering_fields = "__all__"

    @extend_schema(request=None, responses=PerOptionsSerializer)
    def get(self, request, version=None):
        return response.Response(
            PerOptionsSerializer(
                dict(
                    componentratings=PerComponentRating.objects.all().order_by("value"),
                    answers=FormAnswer.objects.all(),
                    overviewassessmenttypes=AssessmentType.objects.all(),
                )
            ).data
        )


class PerProcessStatusViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PerProcessSerializer
    filterset_class = PerOverviewFilter
    permission_classes = [permissions.IsAuthenticated, DenyGuestUserPermission]
    ordering_fields = "__all__"
    get_request_user_regions = RegionRestrictedAdmin.get_request_user_regions
    get_filtered_queryset = RegionRestrictedAdmin.get_filtered_queryset

    def get_queryset(self):
        queryset = Overview.objects.order_by("country", "-assessment_number", "-date_of_assessment")
        return self.get_filtered_queryset(self.request, queryset, dispatch=0)


class PublicPerProcessStatusViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PublicPerProcessSerializer
    filterset_class = PerOverviewFilter
    ordering_fields = "__all__"

    def get_queryset(self):
        return Overview.objects.order_by("country", "-assessment_number", "-date_of_assessment")


class PublicPerProcessStatusViewSet2(PublicPerProcessStatusViewSet):
    """Phase display normalization (mirrors JS)."""

    def list(self, request, *args, **kwargs):
        resp = super().list(request, *args, **kwargs)
        data = resp.data
        results = data.get("results") if isinstance(data, dict) else data
        if results:
            for row in results:
                pd = row.get("phase_display")
                if pd == "Action And Accountability":
                    row["phase_display"] = "Action & accountability"
                elif pd == "WorkPlan":
                    row["phase_display"] = "Workplan"
        return Response(data)


class FormAssessmentViewSet(viewsets.ModelViewSet):
    serializer_class = PerAssessmentSerializer
    permission_classes = [permissions.IsAuthenticated, PerGeneralPermission, DenyGuestUserPermission]
    ordering_fields = "__all__"

    def get_queryset(self):
        return PerAssessment.objects.select_related("overview")


class PublicFormAssessmentViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PublicPerAssessmentSerializer
    ordering_fields = "__all__"

    def get_queryset(self):
        return PerAssessment.objects.select_related("overview")


class PublicFormAssessmentViewSet2(PublicFormAssessmentViewSet):
    """Adds affirmative flags & dashboard components (mirrors JS)."""

    def list(self, request, *args, **kwargs):
        resp = super().list(request, *args, **kwargs)
        data = resp.data
        results = data.get("results") if isinstance(data, dict) else data
        if results:
            for assessment in results:
                components = []
                for area in assessment.get("area_responses", []) or []:
                    for comp in area.get("component_responses", []) or []:
                        comp["urban_considerations_simplified"] = _contains_affirmative(comp.get("urban_considerations"))
                        comp["epi_considerations_simplified"] = _contains_affirmative(comp.get("epi_considerations"))
                        comp["migration_considerations_simplified"] = _contains_affirmative(comp.get("migration_considerations"))
                        comp["climate_environmental_considerations_simplified"] = _contains_affirmative(
                            comp.get("climate_environmental_considerations")
                        )
                        cd = comp.get("component_details") or {}
                        rd = comp.get("rating_details") or {}
                        components.append(
                            {
                                "component_id": comp.get("component"),
                                "component_name": cd.get("title") or "",
                                "component_num": cd.get("component_num"),
                                "area_id": cd.get("area"),
                                "area_name": AREA_NAMES.get(cd.get("area")) if isinstance(cd.get("area"), int) else None,
                                "rating_value": rd.get("value") or 0,
                                "rating_title": rd.get("title") or "",
                            }
                        )
                assessment["components"] = components
        return Response(data)


# Consolidated public endpoints (map-data, assessments-processed, dashboard-data)
class PerMapDataView(views.APIView):
    """Public consolidated PER map data.

    Joins latest Overview per country with minimal country info and normalized phase display.
    """

    def get(self, request):
        latest_overviews = (
            Overview.objects.order_by("country_id", "-assessment_number", "-date_of_assessment")
            .distinct("country_id")
            .select_related("country")
        )
        items = []
        for ov in latest_overviews:
            phase_display = ov.get_phase_display() if hasattr(ov, "get_phase_display") else getattr(ov, "phase_display", None)
            if phase_display == "Action And Accountability":
                phase_display = "Action & accountability"
            elif phase_display == "WorkPlan":
                phase_display = "Workplan"
            items.append(
                {
                    "country_id": ov.country_id,
                    "country_name": ov.country.name if ov.country else None,
                    "country_iso3": getattr(ov.country, "iso3", None),
                    "phase": phase_display,
                    "assessment_number": ov.assessment_number,
                    "date_of_assessment": ov.date_of_assessment,
                }
            )
        return Response({"results": items})


class PerAssessmentsProcessedView(views.APIView):
    """Public consolidated PER assessments processed data.

    Flattens component considerations flags and rating info per assessment for downstream use.
    """

    def get(self, request):
        assessments = PerAssessment.objects.select_related("overview", "overview__country").prefetch_related(
            Prefetch(
                "area_responses",
                queryset=AreaResponse.objects.prefetch_related(
                    Prefetch(
                        "component_response",
                        queryset=FormComponentResponse.objects.prefetch_related("question_responses"),
                    )
                ),
            )
        )
        results = []
        for a in assessments:
            components = []
            for ar in a.area_responses.all():
                for cr in ar.component_response.all():
                    cd = getattr(cr, "component_details", None)
                    rd = getattr(cr, "rating_details", None)
                    # When accessed via serializer, details exist; otherwise construct minimally
                    components.append(
                        {
                            "component_id": getattr(cr, "component_id", None),
                            "urban_considerations_simplified": _contains_affirmative(getattr(cr, "urban_considerations", "")),
                            "epi_considerations_simplified": _contains_affirmative(getattr(cr, "epi_considerations", "")),
                            "migration_considerations_simplified": _contains_affirmative(
                                getattr(cr, "migration_considerations", "")
                            ),
                            "climate_environmental_considerations_simplified": _contains_affirmative(
                                getattr(cr, "climate_environmental_considerations", "")
                            ),
                            "component_name": (cd.title if cd else getattr(cr.component, "title", None)),
                            "component_num": (cd.component_num if cd else getattr(cr.component, "component_num", None)),
                            "area_id": (cd.area if cd else getattr(cr.component.area, "id", None)),
                            "area_name": (
                                AREA_NAMES.get(cd.area)
                                if (cd and isinstance(cd.area, int))
                                else getattr(getattr(cr.component, "area", None), "name", None)
                            ),
                            "rating_value": (rd.value if rd else getattr(getattr(cr, "rating", None), "value", None)),
                            "rating_title": (rd.title if rd else getattr(getattr(cr, "rating", None), "title", None)),
                        }
                    )
            results.append(
                {
                    "assessment_id": a.id,
                    "country_id": getattr(a.overview, "country_id", None),
                    "country_name": getattr(getattr(a.overview, "country", None), "name", None),
                    "components": components,
                }
            )
        return Response({"results": results})


class PerDashboardDataView(views.APIView):
    """Public consolidated PER dashboard data.

    Groups latest per country overview and attaches lightweight assessment entries.
    """

    def get(self, request):
        latest_overviews = (
            Overview.objects.order_by("country_id", "-assessment_number", "-date_of_assessment")
            .distinct("country_id")
            .select_related("country")
        )
        # Map assessments by country for quick attach
        assessments_by_country = {}
        for a in PerAssessment.objects.select_related("overview"):
            cid = getattr(a.overview, "country_id", None)
            if cid is None:
                continue
            assessments_by_country.setdefault(cid, []).append(
                {
                    "assessment_id": a.id,
                    "assessment_number": getattr(a.overview, "assessment_number", None),
                    "date_of_assessment": getattr(a.overview, "date_of_assessment", None),
                }
            )
        items = []
        for ov in latest_overviews:
            phase_display = ov.get_phase_display() if hasattr(ov, "get_phase_display") else getattr(ov, "phase_display", None)
            if phase_display == "Action And Accountability":
                phase_display = "Action & accountability"
            elif phase_display == "WorkPlan":
                phase_display = "Workplan"
            items.append(
                {
                    "country_id": ov.country_id,
                    "country_name": ov.country.name if ov.country else None,
                    "country_iso3": getattr(ov.country, "iso3", None),
                    "phase": phase_display,
                    "assessment_number": ov.assessment_number,
                    "date_of_assessment": ov.date_of_assessment,
                    "countryAssessments": sorted(
                        assessments_by_country.get(ov.country_id, []),
                        key=lambda x: (x["date_of_assessment"] or datetime.min),
                        reverse=True,
                    ),
                }
            )
        return Response({"results": items})


class PerFileViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, viewsets.GenericViewSet):
    permission_classes = [permissions.IsAuthenticated, DenyGuestUserPermission]
    serializer_class = PerFileSerializer

    def get_queryset(self):
        if self.request is None:
            return PerFile.objects.none()
        return PerFile.objects.filter(created_by=self.request.user)

    @extend_schema(request=PerFileInputSerializer, responses=PerFileSerializer(many=True))
    @action(
        detail=False,
        url_path="multiple",
        methods=["POST"],
        permission_classes=[permissions.IsAuthenticated, DenyGuestUserPermission],
    )
    def multiple_file(self, request, pk=None, version=None):
        # converts querydict to original dict
        files = [files[0] for files in dict((request.data).lists()).values()]
        data = [{"file": file} for file in files]
        file_serializer = PerFileSerializer(data=data, context={"request": request}, many=True)
        if file_serializer.is_valid():
            file_serializer.save()
            return response.Response(file_serializer.data, status=drf_status.HTTP_201_CREATED)
        return response.Response(file_serializer.errors, status=drf_status.HTTP_400_BAD_REQUEST)


class PerCountryViewSet(viewsets.ReadOnlyModelViewSet):
    def get_serializer_class(self):
        user = self.request.user
        if not user.is_authenticated:
            return PublicPerCountrySerializer
        else:
            return UserPerCountrySerializer

    def get_queryset(self):
        country_id = self.request.GET.get("country_id", None)
        if country_id:
            return (
                Overview.objects.select_related("country", "type_of_assessment")
                .filter(country_id=country_id)
                .order_by("-created_at")[:1]
            )
        return Overview.objects.none()


class PerAggregatedViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PerProcessSerializer
    filterset_class = PerOverviewFilter
    permission_classes = [permissions.IsAuthenticated, DenyGuestUserPermission]
    ordering_fields = ["assessment_number", "phase", "date_of_assessment"]
    get_request_user_regions = RegionRestrictedAdmin.get_request_user_regions
    get_filtered_queryset = RegionRestrictedAdmin.get_filtered_queryset

    def get_queryset(self):
        queryset = Overview.objects.filter(
            id__in=Overview.objects.order_by("country_id", "-assessment_number").distinct("country_id").values("id")
        )
        return self.get_filtered_queryset(self.request, queryset, dispatch=0)


class OpsLearningFilter(filters.FilterSet):
    type_validated = filters.NumberFilter(field_name="type_validated", lookup_expr="exact")
    appeal_document_id = filters.NumberFilter(field_name="appeal_document_id", lookup_expr="exact")
    organization_validated__in = filters.ModelMultipleChoiceFilter(
        label="validated_organizations",
        field_name="organization_validated",
        help_text="Organization GO identifiers, comma separated",
        widget=CSVWidget,
        queryset=OrganizationTypes.objects.all(),
    )
    sector_validated__in = filters.ModelMultipleChoiceFilter(
        label="validated_sectors",
        field_name="sector_validated",
        help_text="Sector identifiers, comma separated",
        widget=CSVWidget,
        queryset=SectorTag.objects.all(),
    )
    per_component_validated__in = filters.ModelMultipleChoiceFilter(
        label="validated_per_components",
        field_name="per_component_validated",
        help_text="PER Component identifiers, comma separated",
        widget=CSVWidget,
        queryset=FormComponent.objects.all(),
    )
    insight_id = filters.NumberFilter(
        label="Base Insight id for used extracts",
        method="get_cache_response",
    )
    insight_sector_id = filters.NumberFilter(label="Sector insight id for used extracts", method="get_cache_response_sector")
    insight_component_id = filters.NumberFilter(
        label="Component insight id for used extracts",
        method="get_cache_response_component",
    )
    # NOTE: overriding the fields for the typing issue
    sector_validated = filters.NumberFilter(field_name="sector_validated", lookup_expr="exact")
    per_component_validated = filters.NumberFilter(field_name="per_component_validated", lookup_expr="exact")
    # NOTE: this field is used in summary generation
    search_extracts = filters.CharFilter(method="get_filter_search_extracts")

    class Meta:
        model = OpsLearning
        fields = {
            "id": ("exact", "in"),
            "created_at": ("exact", "gt", "gte", "lt", "lte"),
            "modified_at": ("exact", "gt", "gte", "lt", "lte"),
            "is_validated": ("exact",),
            "learning": ("exact", "icontains"),
            "learning_validated": ("exact", "icontains"),
            "type_validated": ("exact", "in"),
            "organization_validated": ("exact",),
            "organization_validated__title": ("exact", "in"),
            "appeal_code": ("exact", "in"),
            "appeal_code__code": ("exact", "icontains", "in"),
            "appeal_code__num_beneficiaries": ("exact", "gt", "gte", "lt", "lte"),
            "appeal_code__start_date": ("exact", "gt", "gte", "lt", "lte"),
            "appeal_code__dtype": ("exact", "in"),
            "appeal_code__atype": ("exact", "in"),
            "appeal_code__country": ("exact", "in"),
            "appeal_code__country__name": ("exact", "in"),
            "appeal_code__country__iso": ("exact", "in"),
            "appeal_code__country__iso3": ("exact", "in"),
            "appeal_code__region": ("exact", "in"),
        }

    def get_cache_response(self, queryset, name, value):
        if value and (ops_learning_cache_response := OpsLearningCacheResponse.objects.filter(id=value).first()):
            return queryset.filter(id__in=ops_learning_cache_response.used_ops_learning.all())
        return queryset

    def get_cache_response_sector(self, queryset, name, value):
        if value and (ops_learning_sector_cache_response := OpsLearningSectorCacheResponse.objects.filter(id=value).first()):
            return queryset.filter(id__in=ops_learning_sector_cache_response.used_ops_learning.all())
        return queryset

    def get_cache_response_component(self, queryset, name, value):
        if value and (
            ops_learning_component_cache_response := OpsLearningComponentCacheResponse.objects.filter(id=value).first()
        ):
            return queryset.filter(id__in=ops_learning_component_cache_response.used_ops_learning.all())
        return queryset

    def get_filter_search_extracts(self, queryset, name, value):
        return queryset.filter(
            Q(learning__icontains=value)
            | Q(learning_validated__icontains=value)
            | Q(appeal_code__name__icontains=value)
            | Q(appeal_code__code__icontains=value)
        )


class OpsLearningViewset(viewsets.ModelViewSet):
    """
    A simple ViewSet for viewing and editing OpsLearning records.
    """

    queryset = OpsLearning.objects.all()
    permission_classes = [DenyGuestUserMutationPermission, OpsLearningPermission]
    filterset_class = OpsLearningFilter
    search_fields = (
        "learning",
        "learning_validated",
        "appeal_code__code",
        "appeal_code__name",
        "appeal_code__name_en",
        "appeal_code__name_es",
        "appeal_code__name_fr",
        "appeal_code__name_ar",
    )

    def get_renderers(self):
        serializer = self.get_serializer()
        if isinstance(serializer, OpsLearningCSVSerializer):
            return [NarrowCSVRenderer()]
        return [renderer() for renderer in tuple(api_settings.DEFAULT_RENDERER_CLASSES)]

    def get_queryset(self):
        qs = super().get_queryset()
        if OpsLearning.is_user_admin(self.request.user):
            return qs.select_related(
                "appeal_code",
            ).prefetch_related(
                "sector",
                "sector_validated",
                "organization",
                "organization_validated",
                "per_component",
                "per_component_validated",
                "appeal_code__event__countries_for_preview",
            )
        return (
            qs.filter(is_validated=True)
            .select_related(
                "appeal_code",
            )
            .prefetch_related(
                "sector",
                "sector_validated",
                "organization",
                "organization_validated",
                "per_component",
                "per_component_validated",
                "appeal_code__event__countries_for_preview",
            )
        )

    def get_serializer_class(self):
        if self.request.method == "GET":
            request_format = self.request.GET.get("format", "json")
            if request_format == "csv":
                return OpsLearningCSVSerializer
            elif OpsLearning.is_user_admin(self.request.user):
                return OpsLearningSerializer
            return PublicOpsLearningSerializer
        return OpsLearningInSerializer

    def get_renderer_context(self):
        context = super().get_renderer_context()
        # Force the order from the serializer. Otherwise redundant literal list

        original = [
            "id",
            "appeal_code.code",
            "appeal_code.name",
            "appeal_code.atype",
            "learning",
            "finding",
            "sector",
            "per_component",
            "organization",
            "appeal_code.country",
            "appeal_code.country_name",
            "appeal_code.region",
            "appeal_code.region_name",
            "appeal_code.dtype",
            "appeal_code.start_date",
            "appeal_code.num_beneficiaries",
            "modified_at",
        ]
        displayed = [
            "id",
            "appeal_code",
            "appeal_name",
            "appeal_type",
            "learning",
            "finding",
            "sector",
            "component",
            "organization",
            "country_id",
            "country_name",
            "region_id",
            "region_name",
            "dtype_name",
            "appeal_year",
            "appeal_num_beneficiaries",
            "modified_at",
        ]

        context["header"] = original
        context["labels"] = {i: i for i in context["header"]}
        # We can change the column titles (called "labels"):
        for i, label in enumerate(displayed):
            context["labels"][original[i]] = label
        context["bom"] = True

        return context

    @extend_schema(
        request=None,
        filters=False,
        responses=OpsLearningOrganizationTypeSerializer(many=True),
    )
    @action(
        detail=False,
        methods=["GET"],
        permission_classes=[DenyGuestUserMutationPermission, OpsLearningPermission],
        serializer_class=OpsLearningOrganizationTypeSerializer,
        url_path="organization-type",
    )
    def organization(self, request):
        """
        Get the Organization Types
        """
        queryset = OrganizationTypes.objects.exclude(is_deprecated=True)
        serializer = OpsLearningOrganizationTypeSerializer(queryset, many=True)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = OpsLearningOrganizationTypeSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)

    @extend_schema(
        request=None,
        filters=True,
        responses=OpsLearningSummarySerializer,
    )
    @action(
        detail=False,
        methods=["GET"],
        permission_classes=[DenyGuestUserMutationPermission, OpsLearningPermission],
        url_path="summary",
    )
    def summary(self, request):
        """
        Get the Ops Learning Summary based on the filters
        """
        ops_learning_summary_instance, filter_data = OpslearningSummaryCacheHelper.get_or_create(request, [self.filterset_class])
        if ops_learning_summary_instance.status == OpsLearningCacheResponse.Status.SUCCESS:
            return response.Response(OpsLearningSummarySerializer(ops_learning_summary_instance).data)

        requested_lang = django_get_language()
        transaction.on_commit(
            lambda: generate_summary.delay(
                ops_learning_summary_id=ops_learning_summary_instance.id,
                filter_data=filter_data,
                translation_lazy=requested_lang == "en",
            )
        )
        return response.Response(OpsLearningSummarySerializer(ops_learning_summary_instance).data)

    @extend_schema(
        request=None,
        filters=True,
        responses=OpsLearningStatSerializer,
    )
    @action(
        detail=False,
        methods=["GET"],
        permission_classes=[DenyGuestUserMutationPermission, OpsLearningPermission],
        url_path="stats",
    )
    def stats(self, request):
        """
        Get the Ops Learning stats based on the filters
        """
        queryset = self.filter_queryset(self.get_queryset()).filter(is_validated=True)
        ops_data = queryset.aggregate(
            operations_included=Count("appeal_code", distinct=True),
            learning_extracts=Count("id", distinct=True),
            sector_covered=Count("sector_validated", distinct=True),
            source_used=Count("appeal_document_id", distinct=True),
        )

        learning_by_sector_qs = (
            SectorTag.objects.filter(validated_sectors__in=queryset, title__isnull=False)
            .annotate(sector_id=F("id"), count=Count("validated_sectors", distinct=True))
            .values("sector_id", "title", "count")
        )

        # NOTE: Queryset is unbounded, we may need to add some start_date filter.
        sources_overtime_qs = (
            queryset.filter(appeal_document_id__isnull=False)
            .annotate(
                atype=F("appeal_code__atype"),
                date=F("appeal_code__start_date"),
                count=Count("appeal_document_id", distinct=True),
            )
            .values("atype", "date", "count")
        )

        learning_by_region_qs = (
            Region.objects.filter(appeal__opslearning__in=queryset)
            .annotate(
                region_id=F("id"),
                region_name=F("label"),
                count=Count("appeal__opslearning", distinct=True),
            )
            .values("region_id", "region_name", "count")
        )

        learning_by_country_qs = (
            Country.objects.filter(appeal__opslearning__in=queryset)
            .annotate(
                country_id=F("id"),
                country_name=F("name"),
                count=Count("appeal__opslearning", distinct=True),
            )
            .values("country_id", "country_name", "count")
        )

        data = {
            "operations_included": ops_data["operations_included"],
            "learning_extracts": ops_data["learning_extracts"],
            "sectors_covered": ops_data["sector_covered"],
            "sources_used": ops_data["source_used"],
            "learning_by_region": learning_by_region_qs,
            "learning_by_sector": learning_by_sector_qs,
            "sources_overtime": sources_overtime_qs,
            "learning_by_country": learning_by_country_qs,
        }
        return response.Response(OpsLearningStatSerializer(data).data)


class PerDocumentUploadViewSet(viewsets.ModelViewSet):
    queryset = PerDocumentUpload.objects.all()
    serializer_class = PerDocumentUploadSerializer
    filterset_class = PerDocumentFilter
    permission_classes = [permissions.IsAuthenticated, PerDocumentUploadPermission, DenyGuestUserPermission]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        return filter_per_queryset_by_user_access(user, queryset)
