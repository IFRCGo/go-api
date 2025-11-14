import datetime
import json
from io import BytesIO
from unittest import mock

import pydash
from django.contrib.auth.models import Group, Permission
from django.core import management
from openpyxl import load_workbook

import api.models as models
from api.factories import country, district
from api.factories.event import DisasterTypeFactory, EventFactory
from api.models import Country, Region
from deployments.factories.emergency_project import (
    EmergencyProjectActivityActionFactory,
    EmergencyProjectActivityFactory,
    EmergencyProjectActivitySectorFactory,
    EruFactory,
    ERUOwnerFactory,
    ERUReadinessFactory,
    ERUReadinessTypeFactory,
)
from deployments.factories.project import (
    ProjectFactory,
    SectorFactory,
    SectorTagFactory,
)
from deployments.factories.regional_project import RegionalProjectFactory
from deployments.factories.user import UserFactory
from deployments.models import (
    EmergencyProject,
    EmergencyProjectActivity,
    ERUReadinessType,
    ERUType,
    OperationTypes,
    Personnel,
    ProgrammeTypes,
    Project,
    Statuses,
    VisibilityCharChoices,
)
from main.test_case import APITestCase, SnapshotTestCase

from .factories.personnel import PersonnelDeploymentFactory, PersonnelFactory


class TestProjectAPI(SnapshotTestCase):
    def test_project_list_zero(self):
        # submit list request
        response = self.client.get("/api/v2/project/")

        # check response
        self.assert_200(response)
        self.assertMatchSnapshot(json.loads(response.content))

    def test_project_list_one(self):
        # create instance
        _country = country.CountryFactory(name="country-1", society_name="society-name-1")
        sct = SectorFactory(title="sect-1", order=1)
        sct_1 = SectorTagFactory(title="sec-tag-1", order=2)
        sct_2 = SectorTagFactory(title="sec-tag-2", order=3)
        district1 = district.DistrictFactory(country=_country, name="district-1", code="dct1")
        district2 = district.DistrictFactory(country=_country, name="district-2", code="dct2")
        dtype = DisasterTypeFactory(name="disaster-type-1", summary="summary")
        regional_project = RegionalProjectFactory(name="regional-project-1")
        event = EventFactory(
            countries=[_country.id],
            slug="event-slug",
            districts=[district1.id, district2.id],
            dtype=dtype,
            name="event-1",
        )
        ProjectFactory.create(
            name="project-1",
            project_districts=[district1, district2],
            budget_amount=100000,
            primary_sector=sct,
            event=event,
            regional_project=regional_project,
            secondary_sectors=[sct_1, sct_2],
            dtype=dtype,
            visibility=VisibilityCharChoices.PUBLIC,
            project_country=_country,
            reporting_ns=_country,
            status=Statuses.COMPLETED,
            programme_type=ProgrammeTypes.BILATERAL,
            operation_type=OperationTypes.EMERGENCY_OPERATION,
        )

        # submit list request
        response = self.client.get("/api/v2/project/")

        # check response
        self.assert_200(response)
        self.assertMatchSnapshot(json.loads(response.content))

    def test_project_list_two(self):
        # create instances
        _country = country.CountryFactory(name="country-1", society_name="society-name-1")
        sct = SectorFactory(title="sect-1", order=1)
        sct_1 = SectorTagFactory(title="sec-tag-1", order=2)
        sct_2 = SectorTagFactory(title="sec-tag-2", order=3)
        district1 = district.DistrictFactory(country=_country, name="district-1", code="dct1")
        district2 = district.DistrictFactory(country=_country, name="district-2", code="dct2")
        dtype = DisasterTypeFactory(name="disaster-type-1", summary="summary")
        regional_project = RegionalProjectFactory(name="regional-project-1")
        event = EventFactory(
            countries=[_country.id],
            slug="event-slug",
            districts=[district1.id, district2.id],
            dtype=dtype,
            name="event-1",
        )
        ProjectFactory.create_batch(
            2,
            name="project-1",
            project_districts=[district1, district2],
            budget_amount=100000,
            primary_sector=sct,
            event=event,
            regional_project=regional_project,
            secondary_sectors=[sct_1, sct_2],
            dtype=dtype,
            visibility=VisibilityCharChoices.PUBLIC,
            project_country=_country,
            reporting_ns=_country,
            status=Statuses.COMPLETED,
            programme_type=ProgrammeTypes.BILATERAL,
            operation_type=OperationTypes.EMERGENCY_OPERATION,
        )

        # submit list request
        response = self.client.get("/api/v2/project/")

        # check response
        self.assert_200(response)
        self.assertMatchSnapshot(json.loads(response.content))

    def test_project_create(self):
        # authenticate
        new_user = UserFactory.create()
        self.authenticate(new_user)

        # create project
        new_project_name = "Mock Project for Create API Test"
        new_project = ProjectFactory.stub(
            name=new_project_name,
            visibility=VisibilityCharChoices.PUBLIC,
            user=new_user,
            operation_type=OperationTypes.PROGRAMME,
        )
        new_country = country.CountryFactory()
        new_district = district.DistrictFactory(country=new_country)
        new_sector = SectorFactory()
        new_sector1 = SectorTagFactory()
        new_sector2 = SectorTagFactory()
        new_project = pydash.omit(
            new_project,
            [
                "user",
                "reporting_ns",
                "project_country",
                "event",
                "dtype",
                "regional_project",
            ],
        )
        new_project["reporting_ns"] = new_country.id
        new_project["project_country"] = new_country.id
        new_project["project_districts"] = [new_district.id]
        new_project["primary_sector"] = new_sector.id
        new_project["secondary_sectors"] = [new_sector1.id, new_sector2.id]

        # submit create request
        response = self.client.post("/api/v2/project/", new_project, format="json")

        # check response
        self.assert_201(response)
        self.assertTrue(Project.objects.get(name=new_project_name))

    def test_project_read(self):
        # create instance
        _country = country.CountryFactory(name="country-1", society_name="society-name-1")
        sct = SectorFactory(title="sect-1", order=1)
        sct_1 = SectorTagFactory(title="sec-tag-1", order=2)
        sct_2 = SectorTagFactory(title="sec-tag-2", order=3)
        district1 = district.DistrictFactory(country=_country, name="district-1", code="dct1")
        district2 = district.DistrictFactory(country=_country, name="district-2", code="dct2")
        dtype = DisasterTypeFactory(name="disaster-type-1", summary="summary")
        regional_project = RegionalProjectFactory(name="regional-project-1")
        event = EventFactory(
            countries=[_country.id],
            slug="event-slug",
            districts=[district1.id, district2.id],
            dtype=dtype,
            name="event-1",
        )
        new_project = ProjectFactory.create(
            name="project-1",
            project_districts=[district1, district2],
            budget_amount=100000,
            primary_sector=sct,
            event=event,
            regional_project=regional_project,
            secondary_sectors=[sct_1, sct_2],
            dtype=dtype,
            visibility=VisibilityCharChoices.PUBLIC,
            project_country=_country,
            reporting_ns=_country,
            status=Statuses.COMPLETED,
            programme_type=ProgrammeTypes.BILATERAL,
            operation_type=OperationTypes.EMERGENCY_OPERATION,
        )

        # submit read request
        response = self.client.get(f"/api/v2/project/{new_project.pk}/")

        # check response
        self.assert_200(response)
        self.assertMatchSnapshot(json.loads(response.content))

    def test_project_update(self):
        # create instance
        _country = country.CountryFactory(name="country-1", society_name="society-name-1")
        sct = SectorFactory(title="sect-1", order=1)
        sct_1 = SectorTagFactory(title="sec-tag-1", order=2)
        sct_2 = SectorTagFactory(title="sec-tag-2", order=3)
        district1 = district.DistrictFactory(country=_country, name="district-1", code="dct1")
        district2 = district.DistrictFactory(country=_country, name="district-2", code="dct2")
        dtype = DisasterTypeFactory(name="disaster-type-1", summary="summary")
        regional_project = RegionalProjectFactory(name="regional-project-1")
        event = EventFactory(
            countries=[_country.id],
            slug="event-slug",
            districts=[district1.id, district2.id],
            dtype=dtype,
            name="event-1",
        )
        new_project = ProjectFactory.create(
            name="project-1",
            project_districts=[district1, district2],
            budget_amount=100000,
            primary_sector=sct,
            event=event,
            regional_project=regional_project,
            secondary_sectors=[sct_1, sct_2],
            dtype=dtype,
            visibility=VisibilityCharChoices.PUBLIC,
            project_country=_country,
            reporting_ns=_country,
            status=Statuses.COMPLETED,
            programme_type=ProgrammeTypes.BILATERAL,
            operation_type=OperationTypes.EMERGENCY_OPERATION,
        )

        # authenticate
        self.authenticate()

        new_project = pydash.omit(
            new_project,
            ["_state", "modified_at", "user", "event", "dtype", "regional_project"],
        )
        # update project name
        new_project_name = "Mock Project for Update API Test"
        new_country = country.CountryFactory(
            name="country-2",
            society_name="society-name-2",
        )
        new_district = district.DistrictFactory(
            country=new_country,
            name="district-3",
            code="dct3",
        )
        new_project["name"] = new_project_name
        new_project["reporting_ns"] = new_country.id
        new_project["project_country"] = new_country.id
        new_project["event"] = new_project["event_id"]
        new_project["project_districts"] = [new_district.id]
        new_project["primary_sector"] = sct.id

        # submit update request
        response = self.client.put(f"/api/v2/project/{new_project['id']}/", new_project, format="json")

        # check response
        self.assert_200(response)
        self.assertMatchSnapshot(json.loads(response.content))
        self.assertTrue(Project.objects.get(name=new_project_name))

    def test_project_delete(self):
        # create instance
        sct = SectorFactory()
        new_project = ProjectFactory.create(visibility=VisibilityCharChoices.PUBLIC, primary_sector=sct)

        # authenticate
        self.authenticate()

        # submit delete request
        response = self.client.delete("/api/v2/project/{}/".format(new_project.pk))

        # check response
        self.assert_204(response)
        self.assertMatchSnapshot(response.content)
        self.assertFalse(Project.objects.count())

    def test_personnel_csv_api(self):
        [PersonnelFactory() for i in range(10)]
        # originally it was so rich. TODO: instead of Personnel fields manually create new ones.
        # country_from,country_to,deployment.comments,deployment.country_deployed_to.id,deployment.country_deployed_to.iso,deployment.country_deployed_to.iso3,deployment.country_deployed_to.name,deployment.country_deployed_to.society_name,deployment.event_deployed_to,deployment.id,end_date,id,is_active,molnix_id,molnix_language,molnix_modality,molnix_operation,molnix_region,molnix_role_profile,molnix_scope,molnix_sector,name,role,start_date,type\r
        # ,,,1,gw,wMq,country-OhbVrpoiVgRVIfLBcbfnoGMbJmTPSIAoCLrZaWZkSBvrjnWvgf,society-name-ZcUDIhyfJsONxKmTecQoXsfogyrDOxkxwnQrSRPeMOkIUpkDyr,,1,,1,True,,,,,,,,,,,,\r
        # ,,,2,Eb,NJu,country-VAlmiYIxHGrkqEZsVvZhDejWoRURzZJxfYzaqIhDxRVRqLyOxg,society-name-NoPeODStPAhicctFhgpIiyDxQVSIALVUjAPgFNArcSxnCCpxgR,,2,,2,True,,,,,,,,,,,,\r
        # ,,,3,Os,xNo,country-RWmlNOzBGufzQgliEupaqypCWrvtLUKaqPxSpdQhDtkzRGTXtS,society-name-oiEjDVMxASJEWIZQnWpRWMYfHCHTxeKhdJGmKIjkuHChRnTLFf,,3,,3,True,,,,,,,,,,,,\r
        # ,,,4,TS,Mcn,country-GiOzeLKdBQipsquZzSVuuCroemiXXLgjgkCDuAhIwXnCtDqhfk,society-name-ujfTpwzGdRtqlbzCJVJpgDgZYihadXoimzxROPfLLqebemPCZi,,4,,4,True,,,,,,,,,,,,\r
        # ,,,5,Pe,VRj,country-PNNpcRuyYhyqIUsbHXxGZGCFcsPmuGfgkXIIaOenQOXnRBgnIS,society-name-bDTvcfedlYqJeKoqAyCOzBubyRhIaPUNeWVLcSewGgsYRtMfsW,,5,,5,True,,,,,,,,,,,,\r
        # ,,,6,oB,sYz,country-gMPtcUZKyfXdXvwBAhXoVPMaOXOydtHcuIKjuGSojdRUzCWMKG,society-name-jivfEKVdJzqfzGBXSiWiEJmFzPKmJNVHpperXBuRKfhQABxwmu,,6,,6,True,,,,,,,,,,,,\r
        # ,,,7,Zm,BvZ,country-YUczhnZPQHDRIjVLoecVjFPbENcpSPialOdtYgDNkLeghFMZNo,society-name-VRzuhRKIiuYnWcLlvehrjWAkSxJkCpLcigYONyXkbFfaalfTPL,,7,,7,True,,,,,,,,,,,,\r
        # ,,,8,TH,BXD,country-MRAYuCEViqlRLuZsmfAxlzyKobbJPNOofDmqSkdzNBMqjfxkKh,society-name-RFhoTnmYrsVFyiBnMnsURmAlAYjbsqpNCpxLRPDfaEiuSzRnTy,,8,,8,True,,,,,,,,,,,,\r
        # ,,,9,ax,aew,country-BwmmOmYevmQESSDMSvLKvNtAvkgDYcFoaOoSoDNnpEVXvZVynz,society-name-bJhjsCmOLWxcbmmsloqUlvJplmblqgbRiyPYhvntDpxZxQkHZN,,9,,9,True,,,,,,,,,,,,\r
        # ,,,10,bM,per,country-niMJnLwriUNBeJqqyPkzDfqRBSjIneOUrOSPmTxKQPGMkAjuYB,society-name-hIhRJAevOfxvXrjZoragyoygYhlHUtLZFgHwSKsJrMgdkuWylw,,10,,10,True,,,,,,,,,,,,\r

        url = "/api/v2/personnel/?format=csv"
        # Also unaunthenticated user can use this, but without names:
        # resp = self.client.get(url)
        # self.assert_401(resp)

        self.authenticate()
        resp = self.client.get(url)
        self.assert_200(resp)
        self.assertMatchSnapshot(resp.content.decode("utf-8"))

    def test_project_csv_api(self):
        _country = country.CountryFactory(name="country-1", society_name="society-name-1")
        sct = SectorFactory(title="sect-1", order=1)
        sct_1 = SectorTagFactory(title="sec-tag-1", order=2)
        sct_2 = SectorTagFactory(title="sec-tag-2", order=3)
        district1 = district.DistrictFactory(country=_country, name="district-1", code="dct1")
        district2 = district.DistrictFactory(country=_country, name="district-2", code="dct2")
        dtype = DisasterTypeFactory(name="disaster-type-1", summary="summary")
        regional_project = RegionalProjectFactory(name="regional-project-1")
        event = EventFactory(
            countries=[_country.id],
            slug="event-slug",
            districts=[district1.id, district2.id],
            dtype=dtype,
            name="event-1",
        )
        ProjectFactory.create_batch(
            10,
            name="project-1",
            project_districts=[district1, district2],
            budget_amount=100000,
            primary_sector=sct,
            event=event,
            regional_project=regional_project,
            secondary_sectors=[sct_1, sct_2],
            dtype=dtype,
            visibility=VisibilityCharChoices.PUBLIC,
            project_country=_country,
            reporting_ns=_country,
            status=Statuses.COMPLETED,
            programme_type=ProgrammeTypes.BILATERAL,
            operation_type=OperationTypes.EMERGENCY_OPERATION,
        )

        url = "/api/v2/project/?format=csv"
        resp = self.client.get(url)
        self.assert_200(resp)
        self.assertMatchSnapshot(resp.content.decode("utf-8"))

    def test_global_project_api(self):
        country_1 = country.CountryFactory()
        country_2 = country.CountryFactory()
        ns_1 = country.CountryFactory()
        ns_2 = country.CountryFactory()
        c1_district1 = district.DistrictFactory(country=country_1)
        c1_district2 = district.DistrictFactory(country=country_1)
        c2_district1 = district.DistrictFactory(country=country_2)
        c2_district2 = district.DistrictFactory(country=country_2)
        sct = SectorFactory()
        sct_1 = SectorTagFactory()
        sct_2 = SectorTagFactory()
        sct_3 = SectorTagFactory()
        sct_4 = SectorTagFactory()
        [
            ProjectFactory.create_batch(
                2,
                primary_sector=sct,
                programme_type=programme_type,
                reporting_ns=ns,
                project_districts=project_districts,
                secondary_sectors=secondary_sectors,
                visibility=VisibilityCharChoices.PUBLIC,
                budget_amount=100000,
            )
            for project_districts, secondary_sectors, programme_type in [
                ([c1_district1, c1_district2], [sct_1, sct_2], ProgrammeTypes.BILATERAL),
                ([c1_district1, c1_district2], [sct_3, sct_4], ProgrammeTypes.MULTILATERAL),
                ([c2_district1, c2_district2], [sct_1, sct_3], ProgrammeTypes.DOMESTIC),
                ([c2_district1, c2_district2], [sct_2, sct_4], ProgrammeTypes.DOMESTIC),
            ]
            for ns in [ns_1, ns_2]
        ]

        url = "/api/v2/global-project/overview/"
        resp = self.client.get(url)
        self.assert_200(resp)
        self.assertMatchSnapshot(resp.json())

        url = "/api/v2/global-project/ns-ongoing-projects-stats/"
        resp = self.client.get(url)
        self.assert_200(resp)
        self.assertMatchSnapshot(resp.json())


class TestEmergencyProjectAPI(APITestCase):
    fixtures = ["emergency_project_activity_actions.json"]

    def test_emergency_project(self):
        supplies = {"2": 100, "1": 1, "4": 3}
        EmergencyProjectActivityFactory.create_batch(5, supplies=supplies)
        url = "/api/v2/emergency-project/"
        self.authenticate()
        response = self.client.get(url)
        self.assert_200(response)
        self.assertEqual(len(response.data["results"]), 5)

    def test_emergency_project_create(self):
        old_emergency_project_count = EmergencyProject.objects.count()
        old_emergency_project_activity_count = EmergencyProjectActivity.objects.count()
        country1 = models.Country.objects.create(name="abc")
        country2 = models.Country.objects.create(name="xyz")
        district1 = models.District.objects.create(name="test district1", country=country1)
        district2 = models.District.objects.create(name="test district2", country=country1)
        sector = EmergencyProjectActivitySectorFactory.create()
        action = EmergencyProjectActivityActionFactory.create()
        event = EventFactory.create(countries=[country1.id, country2.id], districts=[district1.id, district2.id])
        reporting_ns = models.Country.objects.create(name="ne")
        deployed_eru = EruFactory.create()
        data = {
            "title": "Emergency title",
            "event": event.id,
            "districts": [district1.id, district2.id],
            "reporting_ns": reporting_ns.id,
            "reporting_ns_contact_email": None,
            "reporting_ns_contact_name": None,
            "reporting_ns_contact_role": None,
            "status": EmergencyProject.ActivityStatus.ON_GOING,
            "activity_lead": EmergencyProject.ActivityLead.NATIONAL_SOCIETY,
            "start_date": "2022-01-01",
            "deployed_eru": deployed_eru.id,
            "country": country1.id,
            "activities": [
                {
                    "sector": sector.id,
                    "action": action.id,
                    "people_count": 2,
                    "male": 3,
                    "female": 5,
                    "people_households": EmergencyProjectActivity.PeopleHouseholds.PEOPLE,
                    "custom_supplies": {"test_supplies": 23, "test_world": 34, "test_emergency": 56},
                    "points": [],
                },
            ],
        }
        url = "/api/v2/emergency-project/"
        self.authenticate()
        response = self.client.post(url, data=data, format="json")
        self.assert_201(response)
        self.assertEqual(EmergencyProject.objects.count(), old_emergency_project_count + 1)
        self.assertEqual(EmergencyProjectActivity.objects.count(), old_emergency_project_activity_count + 1)


class TestERUReadinessAPI(APITestCase):
    def setUp(self):
        super().setUp()
        self.region = Region.objects.create(name=2)
        self.country = Country.objects.create(name="Nepal", iso3="NLP", region=self.region)

        # Create permissions
        management.call_command("make_permissions")

        self.user_admin = UserFactory.create()
        self.country_admin_permission = Permission.objects.filter(codename="country_admin_%s" % self.country.id).first()
        country_group = Group.objects.filter(name="%s Admins" % self.country.name).first()

        self.user_admin.user_permissions.add(self.country_admin_permission)
        self.user_admin.groups.add(country_group)

    def test_eru_readiness_list(self):
        eru_readiness_type_common = {
            "equipment_readiness": ERUReadinessType.ReadinessStatus.READY,
            "people_readiness": ERUReadinessType.ReadinessStatus.READY,
            "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
        }
        eru_readiness_type_1, eru_readiness_type_2 = ERUReadinessTypeFactory.create_batch(
            2,
            type=ERUType.OSH,
            **eru_readiness_type_common,
        )
        eru_owner_1 = ERUOwnerFactory.create(
            national_society_country=self.country,
        )
        ERUReadinessFactory.create(
            eru_owner=eru_owner_1,
            eru_types=[
                eru_readiness_type_1,
                eru_readiness_type_2,
            ],
        )
        eru_owner_2 = ERUOwnerFactory.create(
            national_society_country=self.country,
        )
        ERUReadinessFactory.create(
            eru_owner=eru_owner_2,
            eru_types=[
                eru_readiness_type_1,
                eru_readiness_type_2,
            ],
        )

        url = "/api/v2/eru-readiness/"
        response = self.client.get(url)
        self.assert_200(response)
        self.assertEqual(len(response.data["results"]), 2)

    def test_eru_readiness_create(self):
        eru_owner = ERUOwnerFactory.create(
            national_society_country=self.country,
        )
        eru_readiness_type_common = {
            "equipment_readiness": ERUReadinessType.ReadinessStatus.READY,
            "people_readiness": ERUReadinessType.ReadinessStatus.READY,
            "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
        }
        eru_readiness_type_1 = ERUReadinessTypeFactory.create(
            type=ERUType.OSH,
            **eru_readiness_type_common,
        )
        eru_readiness_type_2 = ERUReadinessTypeFactory.create(
            type=ERUType.CBS,
            **eru_readiness_type_common,
        )

        data = {
            "eru_owner": eru_owner.id,
            "eru_types": [
                {
                    "type": eru_readiness_type_1.type,
                    "equipment_readiness": eru_readiness_type_1.equipment_readiness,
                    "people_readiness": eru_readiness_type_1.people_readiness,
                    "funding_readiness": eru_readiness_type_1.funding_readiness,
                },
                {
                    "type": eru_readiness_type_2.type,
                    "equipment_readiness": eru_readiness_type_2.equipment_readiness,
                    "people_readiness": eru_readiness_type_2.people_readiness,
                    "funding_readiness": eru_readiness_type_2.funding_readiness,
                },
            ],
        }

        url = "/api/v2/eru-readiness/"
        self.authenticate(self.user_admin)
        response = self.client.post(url, data=data, format="json")
        self.assert_201(response)
        self.assertEqual(response.data["eru_owner_details"]["id"], eru_owner.id)
        self.assertEqual(len(response.data["eru_types"]), 2)

    def test_eru_readiness_update(self):
        eru_readiness_type_common = {
            "equipment_readiness": ERUReadinessType.ReadinessStatus.READY,
            "people_readiness": ERUReadinessType.ReadinessStatus.READY,
            "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
        }
        eru_readiness_type_1 = ERUReadinessTypeFactory.create(
            type=ERUType.OSH,
            **eru_readiness_type_common,
        )
        eru_readiness_type_2 = ERUReadinessTypeFactory.create(
            type=ERUType.CBS,
            **eru_readiness_type_common,
        )
        eru_owner = ERUOwnerFactory.create(
            national_society_country=self.country,
        )
        eru_readiness_1 = ERUReadinessFactory.create(eru_owner=eru_owner, eru_types=[eru_readiness_type_1, eru_readiness_type_2])

        # Update ERU readiness and types
        url = f"/api/v2/eru-readiness/{eru_readiness_1.id}/"

        data = {
            "id": eru_readiness_1.id,
            "eru_owner": eru_owner.id,
            "eru_types": [
                {
                    "id": eru_readiness_type_1.id,
                    "type": eru_readiness_type_1.type,
                    "equipment_readiness": ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                    "people_readiness": ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                    "funding_readiness": ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                },
                {
                    "id": eru_readiness_type_2.id,
                    "type": eru_readiness_type_2.type,
                    "equipment_readiness": ERUReadinessType.ReadinessStatus.READY,
                    "people_readiness": ERUReadinessType.ReadinessStatus.READY,
                    "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
                },
                # Add new ERU type
                {
                    "type": ERUType.MHPSS,
                    "equipment_readiness": ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
                    "people_readiness": ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
                    "funding_readiness": ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
                },
            ],
        }

        self.authenticate(self.user_admin)
        response = self.client.patch(url, data=data, format="json")
        self.assert_200(response)
        self.assertEqual(response.data["id"], eru_readiness_1.id)
        self.assertEqual(response.data["eru_owner_details"]["id"], eru_owner.id)
        self.assertEqual(len(response.data["eru_types"]), 3)

        # Check ids of Updated ERU types
        self.assertEqual(
            {
                response.data["eru_types"][0]["id"],
                response.data["eru_types"][1]["id"],
            },
            {
                eru_readiness_type_1.id,
                eru_readiness_type_2.id,
            },
        )

        self.assertEqual(
            {
                response.data["eru_types"][0]["equipment_readiness"],
                response.data["eru_types"][0]["people_readiness"],
                response.data["eru_types"][0]["funding_readiness"],
                response.data["eru_types"][1]["equipment_readiness"],
                response.data["eru_types"][1]["people_readiness"],
                response.data["eru_types"][1]["funding_readiness"],
                response.data["eru_types"][2]["equipment_readiness"],
                response.data["eru_types"][2]["people_readiness"],
                response.data["eru_types"][2]["funding_readiness"],
            },
            {
                ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                ERUReadinessType.ReadinessStatus.NO_CAPACITY,
                ERUReadinessType.ReadinessStatus.READY,
                ERUReadinessType.ReadinessStatus.READY,
                ERUReadinessType.ReadinessStatus.READY,
                ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
                ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
                ERUReadinessType.ReadinessStatus.PARTIAL_CAPACITY,
            },
        )


class AggregatedERUAndRapidResponseViewSetTestCase(APITestCase):

    def setUp(self):
        super().setUp()

        self.event = EventFactory(
            name="Test Event",
            disaster_start_date=datetime.datetime(2025, 3, 1),
        )

        self.country1 = country.CountryFactory(name="Test Country1")
        self.country2 = country.CountryFactory(name="Test Country2")
        self.country3 = country.CountryFactory(name="Test Country3")

        self.eru_owner = ERUOwnerFactory(
            national_society_country=self.country1,
        )

        self.eru = EruFactory.create_batch(
            3,
            event=self.event,
            deployed_to=self.country1,
            eru_owner=self.eru_owner,
            start_date=datetime.datetime(2022, 3, 1),
            end_date=datetime.datetime(2037, 3, 10),
        )

        self.personnel_deployment1 = PersonnelDeploymentFactory(
            event_deployed_to=self.event,
        )

        # Active
        self.personnel_rapid_response = PersonnelFactory(
            type=Personnel.TypeChoices.RR,
            is_active=True,
            country_from=self.country2,
            deployment=self.personnel_deployment1,
            start_date=datetime.datetime(2022, 1, 1),
            end_date=datetime.datetime(2037, 3, 10),
        )

        self.personnel_rapid_response2 = PersonnelFactory(
            type=Personnel.TypeChoices.RR,
            is_active=True,
            country_from=self.country1,
            deployment=self.personnel_deployment1,
            start_date=datetime.datetime(2022, 1, 1),
            end_date=datetime.datetime(2037, 3, 10),
        )
        # Not Active
        self.personnel_rapid_response3 = PersonnelFactory(
            type=Personnel.TypeChoices.RR,
            is_active=False,
            deployment=self.personnel_deployment1,
            start_date=datetime.datetime(2022, 1, 1),
            end_date=datetime.datetime(2023, 3, 10),
        )

    def test_get_aggregated_data(self):
        # mock Timezone
        patcher = mock.patch("django.utils.timezone.now")
        mock_timezone_now = patcher.start()
        mock_timezone_now.return_value.date.return_value = datetime.date(2022, 11, 11)

        url = "/api/v2/aggregated-eru-and-rapid-response/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 1)
        event_data = data["results"][0]
        self.assertEqual(event_data["deployed_personnel_count"], 2)
        self.assertEqual(event_data["deployed_eru_count"], 3)
        patcher.stop()


class ExportERUReadinessViewTest(APITestCase):
    def setUp(self):
        super().setUp()
        self.country1 = Country.objects.create(
            name="Nepal",
            iso3="NLP",
        )
        self.country2 = Country.objects.create(name="India", iso3="IND")
        eru_readiness_type_common1 = {
            "equipment_readiness": ERUReadinessType.ReadinessStatus.READY,
            "people_readiness": ERUReadinessType.ReadinessStatus.NO_CAPACITY,
            "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
        }
        eru_readiness_type_common2 = {
            "equipment_readiness": ERUReadinessType.ReadinessStatus.NO_CAPACITY,
            "people_readiness": ERUReadinessType.ReadinessStatus.READY,
            "funding_readiness": ERUReadinessType.ReadinessStatus.READY,
        }
        eru_readiness_type_1 = ERUReadinessTypeFactory.create(
            type=ERUType.MHPSS,
            **eru_readiness_type_common1,
            comment="Test comment1",
        )
        eru_readiness_type_2 = ERUReadinessTypeFactory.create(
            type=ERUType.TELECOM,
            **eru_readiness_type_common1,
            comment="Test comment2",
        )
        eru_readiness_type_3 = ERUReadinessTypeFactory.create(
            type=ERUType.MHPSS,
            **eru_readiness_type_common2,
            comment="Test comment3",
        )
        eru_owner1 = ERUOwnerFactory.create(
            national_society_country=self.country1,
        )
        eru_owner2 = ERUOwnerFactory.create(
            national_society_country=self.country2,
        )
        ERUReadinessFactory.create(eru_owner=eru_owner1, eru_types=[eru_readiness_type_1, eru_readiness_type_2])
        ERUReadinessFactory.create(eru_owner=eru_owner2, eru_types=[eru_readiness_type_3])

    def test_export_response(self):
        url = "/api/v2/export-eru-readiness/"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        with BytesIO(response.content) as excel_file:
            wb = load_workbook(filename=excel_file)
            ws = wb.active

            self.assertEqual(ws["A1"].value, "National Society")
            self.assertEqual(ws["B1"].value, "Updated Date")
            self.assertEqual(ws["C1"].value, "IT & Telecom")
            self.assertEqual(ws["G1"].value, "Logistics")

            # Check sub-headers
            self.assertEqual(ws["C2"].value, "Equipment")
            self.assertEqual(ws["D2"].value, "People")
            self.assertEqual(ws["E2"].value, "Funding")
            self.assertEqual(ws["F2"].value, "Comment")

            row1 = list(ws.iter_rows(min_row=3))[0]
            self.assertEqual(row1[0].value, "Nepal")
            self.assertEqual(row1[2].value, "Ready")
            self.assertEqual(row1[3].value, "No capacity")
            self.assertEqual(row1[4].value, "Ready")
            self.assertEqual(row1[5].value, "Test comment2")

            row2 = list(ws.iter_rows(min_row=4))[0]
            self.assertEqual(row2[0].value, "India")
